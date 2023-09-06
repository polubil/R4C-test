import os
from datetime import datetime, timedelta

from django.test import TestCase
from django.urls import reverse
from django.utils import timezone
from openpyxl import load_workbook

from R4C import settings
from robots.models import Robot
from robots.tasks import generate_report

START_DATE = datetime.now(tz=timezone.utc) - timedelta(days=7)
END_DATE = datetime.now(tz=timezone.utc)


class RobotTestCase(TestCase):
    def setUp(self):
        self.start_date = START_DATE
        self.end_date = END_DATE

        robot = Robot.objects.create(serial="A1-B1", model="A1", version="B1",
                                     created=self.end_date-timedelta(days=1))
        robot.save()

    def test_single_robot_creation(self):
        robot = Robot.objects.filter(serial="A1-B1")
        self.assertTrue(robot.exists())
        self.assertEqual(robot.first().model, "A1")
        self.assertEqual(robot.first().version, "B1")
        self.assertTrue(isinstance(robot.first().created, datetime))

    def test_single_robot_report_generation(self):

        generate_report(self.start_date, self.end_date)

        filename = f"{self.start_date.strftime('%Y.%m.%d')}-{self.end_date.strftime('%Y.%m.%d')}.xlsx"
        path_to_file = f"{settings.REPORT_DIR}/{filename}"

        self.assertTrue(os.path.exists(path_to_file))

        wb = load_workbook(filename=path_to_file)
        self.assertEqual(wb.sheetnames, ["A1"])

        sheet = wb["A1"]
        self.assertEqual(
            [sheet['A1'].value, sheet['B1'].value, sheet['C1'].value],
            ["Модель", "Версия", "Количество за месяц"]
        )
        self.assertEqual(
            [sheet['A2'].value, sheet['B2'].value, sheet['C2'].value],
            ["A1", "B1", 1]
        )

    def test_single_robot_report_view(self):
        url = reverse('report')
        response = self.client.get(url)
        self.assertEqual(response.status_code, 200)
        self.assertTemplateUsed(response, 'robots/templates/report.html')

    def test_single_robot_report_exists(self):

        start_date_str = self.start_date.strftime('%Y.%m.%d')
        end_date_str = self.end_date.strftime('%Y.%m.%d')
        filename = f"{start_date_str}-{end_date_str}.xlsx"
        url = reverse("download_report", args=[filename])
        response = self.client.get(url)
        self.assertEqual(response.status_code, 200)


class ReportMultipleTestCase(TestCase):

    def setUp(self):
        self.start_date = START_DATE
        self.end_date = END_DATE
        robot1 = Robot.objects.create(serial="A1-B1", model="A1", version="B1",
                                      created=self.end_date - timedelta(days=1))
        robot2 = Robot.objects.create(serial="A2-B1", model="A2", version="B1",
                                      created=self.end_date - timedelta(days=2))
        robot3 = Robot.objects.create(serial="A1-B2", model="A1", version="B2",
                                      created=self.end_date - timedelta(days=3))
        robot1.save()
        robot2.save()
        robot3.save()

    def test_multiple_robots_creation(self):
        robot1 = Robot.objects.filter(serial="A1-B1")
        robot2 = Robot.objects.filter(serial="A2-B1")
        robot3 = Robot.objects.filter(serial="A1-B2")
        self.assertTrue(robot1.exists())
        self.assertTrue(robot2.exists())
        self.assertTrue(robot3.exists())
        self.assertEqual(robot1.first().model, "A1")
        self.assertEqual(robot2.first().model, "A2")
        self.assertEqual(robot3.first().model, "A1")
        self.assertEqual(robot1.first().version, "B1")
        self.assertEqual(robot2.first().version, "B1")
        self.assertEqual(robot3.first().version, "B2")
        self.assertTrue(isinstance(robot1.first().created, datetime))
        self.assertTrue(isinstance(robot2.first().created, datetime))
        self.assertTrue(isinstance(robot3.first().created, datetime))
        self.assertEqual(Robot.objects.all().count(), 3)

    def test_multiple_robots_report_generation(self):
        generate_report(self.start_date, self.end_date)

        start_date_str = self.start_date.strftime('%Y.%m.%d')
        end_date_str = self.end_date.strftime('%Y.%m.%d')

        filename = f"{start_date_str}-{end_date_str}.xlsx"
        path_to_file = f"{settings.REPORT_DIR}/{filename}"

        self.assertTrue(os.path.exists(path_to_file))

        wb = load_workbook(filename=path_to_file)
        self.assertEqual(set(wb.sheetnames), {"A1", "A2"})

        sheet = wb["A1"]
        self.assertEqual(
            [sheet['A1'].value, sheet['B1'].value, sheet['C1'].value],
            ["Модель", "Версия", "Количество за месяц"]
        )
        self.assertEqual(
            [sheet['A2'].value, sheet['B2'].value, sheet['C2'].value],
            ["A1", "B1", 1]
        )

        sheet2 = wb.get_sheet_by_name("A2")
        self.assertEqual(
            [sheet2['A1'].value, sheet2['B1'].value, sheet2['C1'].value],
            ["Модель", "Версия", "Количество за месяц"]
        )
        self.assertEqual(
            [sheet2['A2'].value, sheet2['B2'].value, sheet2['C2'].value],
            ["A2", "B1", 1]
        )

    def test_multiple_robots_report_view(self):
        url = reverse('report')
        response = self.client.get(url)
        self.assertEqual(response.status_code, 200)
        self.assertTemplateUsed(response, 'robots/templates/report.html')

    def test_multiple_robots_report_exists(self):

        generate_report(self.start_date, self.end_date)

        start_date_str = self.start_date.strftime('%Y.%m.%d')
        end_date_str = self.end_date.strftime('%Y.%m.%d')
        filename = f"{start_date_str}-{end_date_str}.xlsx"
        url = reverse("download_report", args=[filename])
        response = self.client.get(url)
        self.assertEqual(response.status_code, 200)
